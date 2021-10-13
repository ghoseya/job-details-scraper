from PageExtractor import dataExtraction, field_entities
from openpyxl import Workbook, load_workbook
from os.path import exists
from os import makedirs
from sys import exit
from time import sleep
from pandas import read_excel
from pickle import load


def main(file_path: str):
    with open("job_urls.txt", "rb") as fp:
        primary_gen_link = list(set(load(fp))) # "set" to have unique elements & "list" to have ordered continuous data 
    if exists(file_path):
        read_data = read_excel(file_path)
        read_data_jobs = read_data["jobs_url"].tolist()
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        workbook.save(filename=file_path)
    else:
        read_data_jobs = []
        workbook = Workbook()
        workbook.save(filename=file_path)
        sheet = workbook.active
        temp_list = field_entities() + ["jobs_url"]
        sheet.append(temp_list)
        workbook.save(filename=file_path)


    for jobs in primary_gen_link:
        print(jobs + " - Processing")

        if (jobs in read_data_jobs):
            print("Data Present")
        else:
            params_data = dataExtraction(jobs)
            sheet.append(params_data + [jobs])
            workbook.save(filename=file_path)
            print(jobs + " - Completed")
            sleep(3)

    print("program completed successfully")
    exit()

if __name__ == "__main__":
    if exists("datasets"):
        pass
    else:
        makedirs("datasets")
    file_name = input('Enter filename without extension: ')+'.xlsx'
    rel_file_path = 'datasets/'+file_name
    while True:
        try:
            main(rel_file_path)
        except AttributeError as e:
            print(e)
            sleep(7)
            main(rel_file_path)
